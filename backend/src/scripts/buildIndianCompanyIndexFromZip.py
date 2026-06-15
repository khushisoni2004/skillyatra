import csv
import json
import re
import zipfile
from collections import Counter
from io import BytesIO, TextIOWrapper
from pathlib import Path

DATASET_DIR = Path("src/data/datasets")
ZIP_PATH = DATASET_DIR / "Indian Job Market Dataset 2025 97k Data Points.zip"
OUT_FILE = Path("src/data/companyRoleIndex.json")

TOP_COMPANY_LIMIT = 2000
MAX_ROLES_PER_COMPANY = 25
MAX_SKILLS_PER_ROLE = 12
MAX_LOCATIONS_PER_ROLE = 6
MAX_KEYWORDS_PER_ROLE = 10
MAX_SALARY_SAMPLES = 5

TECH_SKILLS = [
    "python", "java", "javascript", "typescript", "react", "node", "express",
    "mongodb", "mysql", "sql", "postgresql", "html", "css", "tailwind",
    "c++", "c#", "spring", "spring boot", "django", "flask", "fastapi",
    "aws", "azure", "gcp", "docker", "kubernetes", "git", "github",
    "machine learning", "deep learning", "data science", "data analysis",
    "excel", "power bi", "tableau", "pandas", "numpy", "scikit", "tensorflow",
    "pytorch", "nlp", "api", "rest", "microservices", "devops", "linux",
    "testing", "selenium", "manual testing", "automation testing", "jira",
    "agile", "scrum", "figma", "ui", "ux", "android", "ios", "flutter",
    "kotlin", "swift", "php", "laravel", "wordpress", "sap", "salesforce",
    "cybersecurity", "networking", "cloud", "data structures", "dsa", "oops",
    "dbms", "operating system", "computer networks"
]

ROLE_KEYWORDS = [
    "software", "developer", "engineer", "frontend", "backend", "full stack",
    "data", "analyst", "business analyst", "qa", "tester", "devops", "cloud",
    "network", "support", "consultant", "manager", "intern", "trainee",
    "associate", "architect", "administrator", "designer"
]

def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()

def norm(value):
    return clean_text(value).lower()

def find_col(columns, candidates):
    clean_map = {clean_text(c).lower(): c for c in columns}

    for cand in candidates:
        key = cand.lower()
        if key in clean_map:
            return clean_map[key]

    for c in columns:
        cl = clean_text(c).lower()
        for cand in candidates:
            if cand.lower() in cl:
                return c

    return None

def split_skills(value):
    text = clean_text(value)
    if not text:
        return []

    parts = re.split(r"[,|;/•]+", text)
    skills = []

    for p in parts:
        s = clean_text(p)
        if 1 < len(s) <= 50:
            skills.append(s)

    return list(dict.fromkeys(skills))[:30]

def extract_skills_from_text(*values):
    joined = " ".join(clean_text(v) for v in values).lower()
    found = []

    for skill in TECH_SKILLS:
        if skill in joined:
            if skill in ["sql", "aws", "gcp", "api", "ui", "ux", "php", "nlp", "dsa", "oops", "dbms"]:
                found.append(skill.upper())
            else:
                found.append(skill.title())

    return list(dict.fromkeys(found))

def extract_keywords(*values):
    joined = " ".join(clean_text(v) for v in values).lower()
    found = []

    for kw in ROLE_KEYWORDS:
        if kw in joined:
            found.append(kw.title())

    return list(dict.fromkeys(found))

def build_prepare_plan(role_name, skills):
    role = norm(role_name)
    focus = skills[:8]
    plan = []

    if any(x in role for x in ["software", "developer", "engineer", "full stack", "frontend", "backend"]):
        plan.append("Practice DSA daily: arrays, strings, linked list, stack, queue, trees, graphs and basic DP.")
        plan.append("Revise CS fundamentals: OOP, DBMS, OS, CN and SQL.")
        plan.append("Build one project matching this role and add GitHub + deployment link.")
    elif "data" in role or "analyst" in role:
        plan.append("Prepare SQL, Excel, statistics, Python and data visualization.")
        plan.append("Practice joins, group by, window functions and dashboard case studies.")
        plan.append("Build one analytics project using a real dataset.")
    elif "qa" in role or "test" in role:
        plan.append("Prepare manual testing, test cases, bug reports, SDLC, STLC and basic automation.")
        plan.append("Learn Selenium or API testing basics.")
        plan.append("Create one testing project with clear test cases.")
    elif "cloud" in role or "devops" in role:
        plan.append("Prepare Linux, Git, Docker, CI/CD, AWS basics and deployment flow.")
        plan.append("Practice basic cloud architecture and troubleshooting.")
        plan.append("Deploy one full-stack project and document the pipeline.")
    elif "network" in role:
        plan.append("Prepare OSI model, TCP/IP, subnetting, routing, DNS, DHCP and firewall basics.")
        plan.append("Practice network troubleshooting scenarios.")
        plan.append("Learn basic cloud networking: VPC, subnet, security group and load balancer.")
    elif "business analyst" in role or "manager" in role:
        plan.append("Prepare requirement gathering, SDLC, Agile, Jira, user stories and documentation.")
        plan.append("Practice case studies and communication-based interview questions.")
        plan.append("Create one sample BRD/FRD or project workflow document.")
    else:
        plan.append("Prepare the listed role skills and add resume proof for each important skill.")
        plan.append("Revise aptitude, communication, SQL and basic CS fundamentals.")
        plan.append("Create one project or practice proof related to this role.")

    if focus:
        plan.append("Priority skills from dataset: " + ", ".join(focus) + ".")

    return plan

def supported_files(zip_file):
    files = []
    for name in zip_file.namelist():
        low = name.lower()
        if low.endswith(".csv") or low.endswith(".xlsx"):
            files.append(name)
    return files

def choose_file(zip_file):
    files = supported_files(zip_file)

    if not files:
        print("\nNo CSV/XLSX found. Files inside ZIP:")
        for name in zip_file.namelist():
            print("-", name)
        raise RuntimeError("No supported data file found inside ZIP. Need CSV or XLSX.")

    best = files[0]
    best_score = -1

    for name in files:
        low = name.lower()
        score = 0

        for word in ["job", "market", "india", "company", "data"]:
            if word in low:
                score += 2

        if low.endswith(".csv"):
            score += 1
        if low.endswith(".xlsx"):
            score += 1

        if score > best_score:
            best_score = score
            best = name

    return best

def rows_from_csv(zip_file, name):
    with zip_file.open(name) as f:
        wrapper = TextIOWrapper(f, encoding="utf-8", errors="ignore")
        reader = csv.DictReader(wrapper)
        for row in reader:
            yield row

def rows_from_xlsx(zip_file, name):
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl not installed. Run: python3 -m pip install --user openpyxl")

    data = zip_file.read(name)
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return

    columns = [clean_text(h) for h in headers]

    for values in rows:
        row = {}
        for col, val in zip(columns, values):
            row[col] = val
        yield row

def get_columns_from_first_row(row):
    return list(row.keys())

def main():
    if not ZIP_PATH.exists():
        raise FileNotFoundError(f"Dataset ZIP not found: {ZIP_PATH}")

    print("Using ZIP:", ZIP_PATH)

    companies = {}
    total_rows = 0

    with zipfile.ZipFile(ZIP_PATH) as z:
        data_file = choose_file(z)
        print("Using data file:", data_file)

        if data_file.lower().endswith(".csv"):
            iterator = rows_from_csv(z, data_file)
        elif data_file.lower().endswith(".xlsx"):
            iterator = rows_from_xlsx(z, data_file)
        else:
            raise RuntimeError("Unsupported file type.")

        first_row = next(iterator, None)

        if not first_row:
            raise RuntimeError("Dataset file is empty.")

        columns = get_columns_from_first_row(first_row)

        print("\nColumns found:")
        for c in columns:
            print("-", c)

        company_col = find_col(columns, ["company", "company_name", "company name", "employer", "organization"])
        title_col = find_col(columns, ["job_title", "job title", "title", "role", "position", "designation"])
        location_col = find_col(columns, ["location", "job_location", "job location", "city", "state"])
        skills_col = find_col(columns, ["skills", "key_skills", "key skills", "required_skills", "required skills"])
        exp_col = find_col(columns, ["experience", "experience_required", "exp", "experience range"])
        salary_col = find_col(columns, ["salary", "pay", "ctc", "package", "salary range"])
        desc_col = find_col(columns, ["description", "job_description", "job description", "details", "summary"])

        if not company_col or not title_col:
            raise RuntimeError(f"Missing company/title columns. Available columns: {columns}")

        print("\nDetected mapping:")
        print("company:", company_col)
        print("title:", title_col)
        print("location:", location_col)
        print("skills:", skills_col)
        print("experience:", exp_col)
        print("salary:", salary_col)
        print("description:", desc_col)

        all_rows = [first_row]
        all_rows.extend(iterator)

        for row in all_rows:
            total_rows += 1

            company = clean_text(row.get(company_col))
            title = clean_text(row.get(title_col))

            if not company or not title:
                continue

            location = clean_text(row.get(location_col)) if location_col else ""
            salary = clean_text(row.get(salary_col)) if salary_col else ""
            exp = clean_text(row.get(exp_col)) if exp_col else ""
            desc = clean_text(row.get(desc_col)) if desc_col else ""

            raw_skills = split_skills(row.get(skills_col)) if skills_col else []
            detected_skills = extract_skills_from_text(title, desc, " ".join(raw_skills))
            skills = list(dict.fromkeys(raw_skills + detected_skills))
            keywords = extract_keywords(title, desc)

            if company not in companies:
                companies[company] = {
                    "companyName": company,
                    "totalJobCount": 0,
                    "roleMap": {}
                }

            companies[company]["totalJobCount"] += 1

            role_key = title.lower().strip()
            role_map = companies[company]["roleMap"]

            if role_key not in role_map:
                role_map[role_key] = {
                    "roleName": title,
                    "jobCount": 0,
                    "skillCounter": Counter(),
                    "locationCounter": Counter(),
                    "keywordCounter": Counter(),
                    "experienceCounter": Counter(),
                    "salarySamples": []
                }

            role = role_map[role_key]
            role["jobCount"] += 1

            for s in skills:
                role["skillCounter"][s] += 1

            if location:
                role["locationCounter"][location] += 1

            if exp:
                role["experienceCounter"][exp] += 1

            if salary and len(role["salarySamples"]) < MAX_SALARY_SAMPLES:
                role["salarySamples"].append(salary)

            for kw in keywords:
                role["keywordCounter"][kw] += 1

    company_list = sorted(
        companies.values(),
        key=lambda c: c["totalJobCount"],
        reverse=True
    )[:TOP_COMPANY_LIMIT]

    final_companies = []

    for c in company_list:
        roles_sorted = sorted(
            c["roleMap"].values(),
            key=lambda r: r["jobCount"],
            reverse=True
        )[:MAX_ROLES_PER_COMPANY]

        final_roles = []

        for r in roles_sorted:
            skills = [x for x, _ in r["skillCounter"].most_common(MAX_SKILLS_PER_ROLE)]
            keywords = [x for x, _ in r["keywordCounter"].most_common(MAX_KEYWORDS_PER_ROLE)]
            locations = [x for x, _ in r["locationCounter"].most_common(MAX_LOCATIONS_PER_ROLE)]
            experience = [x for x, _ in r["experienceCounter"].most_common(5)]

            final_roles.append({
                "roleName": r["roleName"],
                "jobCount": r["jobCount"],
                "skills": skills,
                "requiredSkills": skills,
                "preparationFocus": skills if skills else keywords,
                "commonKeywords": keywords,
                "locations": locations,
                "experience": experience,
                "salarySamples": r["salarySamples"],
                "howToPrepare": build_prepare_plan(r["roleName"], skills if skills else keywords)
            })

        final_companies.append({
            "companyName": c["companyName"],
            "totalJobCount": c["totalJobCount"],
            "roles": final_roles
        })

    output = {
        "source": "Indian Job Market Dataset 2025 97k Data Points.zip",
        "dataFileUsed": data_file,
        "totalCompanies": len(final_companies),
        "totalJobRows": total_rows,
        "companies": final_companies,
        "disclaimer": "Data is generated only from the uploaded Indian job market dataset. Students should verify latest official company career pages before applying."
    }

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    tmp = OUT_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(output, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    tmp.replace(OUT_FILE)

    print("\nIndex created:", OUT_FILE)
    print("Total companies:", output["totalCompanies"])
    print("Total job rows:", output["totalJobRows"])
    print("Index size:", round(OUT_FILE.stat().st_size / 1024 / 1024, 2), "MB")
    print("Top companies:")
    for c in final_companies[:20]:
        print("-", c["companyName"], c["totalJobCount"])

if __name__ == "__main__":
    main()
