import json
import zipfile
from io import BytesIO
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

ZIP_PATH = Path("src/data/datasets/Indian Job Market Dataset 2025 97k Data Points.zip")
INDEX_PATH = Path("src/data/companyRoleIndex.json")

def clean(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", " ").strip()

def num(v):
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None

def split_skills(v):
    text = clean(v)
    if not text:
        return []
    parts = []
    for sep in [",", "|", ";", "/", "•"]:
        text = text.replace(sep, ",")
    for p in text.split(","):
        s = clean(p)
        if 1 < len(s) <= 60:
            parts.append(s)
    return parts

def top(counter, n):
    return [{"name": k, "count": v} for k, v in counter.most_common(n)]

if not ZIP_PATH.exists():
    raise FileNotFoundError(f"Missing dataset ZIP: {ZIP_PATH}")

if not INDEX_PATH.exists():
    raise FileNotFoundError(f"Missing company index: {INDEX_PATH}")

index = json.loads(INDEX_PATH.read_text())

companies = index.get("companies", [])
company_names = {c["companyName"].lower(): c["companyName"] for c in companies}

stats = defaultdict(lambda: {
    "locations": Counter(),
    "skills": Counter(),
    "roles": Counter(),
    "jobUploads": Counter(),
    "ratings": [],
    "reviews": [],
    "minSalary": [],
    "maxSalary": [],
    "minExperience": [],
    "maxExperience": [],
    "salaryText": Counter(),
    "experienceText": Counter()
})

with zipfile.ZipFile(ZIP_PATH) as z:
    xlsx_files = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
    if not xlsx_files:
        raise RuntimeError("No XLSX file found inside Indian dataset ZIP.")

    xlsx_name = xlsx_files[0]
    print("Reading:", xlsx_name)

    wb = load_workbook(BytesIO(z.read(xlsx_name)), read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    col = {h: i for i, h in enumerate(headers)}

    required = ["companyName", "title"]
    for r in required:
        if r not in col:
            raise RuntimeError(f"Required column missing: {r}")

    total = 0
    matched = 0

    for row in rows:
        total += 1

        company = clean(row[col["companyName"]])
        if not company:
            continue

        key = company.lower()
        if key not in company_names:
            continue

        matched += 1
        canonical = company_names[key]
        s = stats[canonical]

        title = clean(row[col["title"]]) if "title" in col else ""
        location = clean(row[col["location"]]) if "location" in col else ""
        skills = clean(row[col["tagsAndSkills"]]) if "tagsAndSkills" in col else ""
        uploaded = clean(row[col["jobUploaded"]]) if "jobUploaded" in col else ""
        rating = num(row[col["AggregateRating"]]) if "AggregateRating" in col else None
        reviews = num(row[col["ReviewsCount"]]) if "ReviewsCount" in col else None
        min_sal = num(row[col["minimumSalary"]]) if "minimumSalary" in col else None
        max_sal = num(row[col["maximumSalary"]]) if "maximumSalary" in col else None
        min_exp = num(row[col["minimumExperience"]]) if "minimumExperience" in col else None
        max_exp = num(row[col["maximumExperience"]]) if "maximumExperience" in col else None
        salary_text = clean(row[col["salary"]]) if "salary" in col else ""
        exp_text = clean(row[col["experience"]]) if "experience" in col else ""

        if title:
            s["roles"][title] += 1

        if location:
            s["locations"][location] += 1

        for skill in split_skills(skills):
            s["skills"][skill] += 1

        if uploaded:
            s["jobUploads"][uploaded] += 1

        if rating is not None and rating > 0:
            s["ratings"].append(rating)

        if reviews is not None and reviews >= 0:
            s["reviews"].append(reviews)

        if min_sal is not None and min_sal > 0:
            s["minSalary"].append(min_sal)

        if max_sal is not None and max_sal > 0:
            s["maxSalary"].append(max_sal)

        if min_exp is not None and min_exp >= 0:
            s["minExperience"].append(min_exp)

        if max_exp is not None and max_exp >= 0:
            s["maxExperience"].append(max_exp)

        if salary_text:
            s["salaryText"][salary_text] += 1

        if exp_text:
            s["experienceText"][exp_text] += 1

print("Rows scanned:", total)
print("Rows matched with indexed companies:", matched)

for c in companies:
    name = c["companyName"]
    s = stats.get(name)

    if not s:
        c["companyInfo"] = {
            "source": "Indian Job Market Dataset 2025",
            "note": "Company exists in index, but detailed company info was not available in matched dataset rows."
        }
        continue

    ratings = s["ratings"]
    reviews = s["reviews"]

    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else None
    max_reviews = int(max(reviews)) if reviews else None

    min_salary = int(min(s["minSalary"])) if s["minSalary"] else None
    max_salary = int(max(s["maxSalary"])) if s["maxSalary"] else None

    min_exp = int(min(s["minExperience"])) if s["minExperience"] else None
    max_exp = int(max(s["maxExperience"])) if s["maxExperience"] else None

    c["companyInfo"] = {
        "source": "Indian Job Market Dataset 2025",
        "totalJobsInDataset": c.get("totalJobCount", 0),
        "topLocations": top(s["locations"], 8),
        "topSkills": top(s["skills"], 15),
        "topRoles": top(s["roles"], 10),
        "recentJobUploads": top(s["jobUploads"], 5),
        "rating": avg_rating,
        "reviewsCount": max_reviews,
        "salaryRange": {
            "minimumSalary": min_salary,
            "maximumSalary": max_salary,
            "commonSalaryTexts": top(s["salaryText"], 5)
        },
        "experienceRange": {
            "minimumExperience": min_exp,
            "maximumExperience": max_exp,
            "commonExperienceTexts": top(s["experienceText"], 5)
        }
    }

index["disclaimer"] = "Company, role, skill, salary, experience, rating and location information is generated only from the uploaded Indian Job Market Dataset 2025. Students should verify latest official company career pages before applying."

tmp = INDEX_PATH.with_suffix(".json.tmp")
tmp.write_text(json.dumps(index, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
tmp.replace(INDEX_PATH)

print("Enriched index saved:", INDEX_PATH)
print("Index size MB:", round(INDEX_PATH.stat().st_size / 1024 / 1024, 2))
